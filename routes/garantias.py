import io
import json
import logging
import os
import uuid
from datetime import date, datetime

from flask import Blueprint, jsonify, request, send_file, redirect
from services.s3_service import subir_archivo_s3, generar_url_firmada_s3, existe_archivo_s3
from werkzeug.utils import secure_filename

from db_conexion import obtener_conexion
from services.garantias_service import exportar_excel, get_dashboard_data, invalidar_cache
from utils.jwt_utils import verificar_token

garantias_bp = Blueprint("garantias", __name__, url_prefix="/garantias")

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'uploads', 'garantias')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif', 'mp4', 'mov', 'avi', 'webp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Dashboard (existing) ─────────────────────────────────────────────────────

@garantias_bp.route("/dashboard", methods=["GET"])
def dashboard():
    try:
        data = get_dashboard_data()
        return jsonify(data)
    except Exception as e:
        logging.exception("Error en /garantias/dashboard: %s", e)
        return jsonify({"error": "Error al obtener datos de garantias"}), 500


@garantias_bp.route("/exportar", methods=["GET"])
def exportar():
    try:
        excel_bytes = exportar_excel()
        buf = io.BytesIO(excel_bytes)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="garantias.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        logging.exception("Error al exportar garantias Excel: %s", e)
        return jsonify({"error": "Error al generar Excel"}), 500


@garantias_bp.route("/refrescar", methods=["POST"])
def refrescar():
    invalidar_cache()
    return jsonify({"ok": True, "mensaje": "Cache invalidado"})


# ── DB initialization ─────────────────────────────────────────────────────────

@garantias_bp.route("/inicializar-tablas", methods=["POST"])
def inicializar_tablas():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS garantia_formularios (
                id INT AUTO_INCREMENT PRIMARY KEY,
                folio VARCHAR(50) UNIQUE,
                email VARCHAR(255),
                distribuidor VARCHAR(255),
                contacto VARCHAR(255),
                puesto VARCHAR(255),
                marca VARCHAR(100),
                datos LONGTEXT,
                estatus VARCHAR(50) DEFAULT 'Enviado',
                fecha_creacion DATETIME DEFAULT NOW(),
                fecha_actualizacion DATETIME DEFAULT NOW() ON UPDATE NOW()
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS garantia_estructura (
                id INT AUTO_INCREMENT PRIMARY KEY,
                version INT DEFAULT 1,
                estructura LONGTEXT,
                fecha_creacion DATETIME DEFAULT NOW()
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS garantia_comentarios (
                id INT AUTO_INCREMENT PRIMARY KEY,
                formulario_id INT NOT NULL,
                autor VARCHAR(120) NOT NULL DEFAULT 'Sistema',
                texto TEXT NOT NULL,
                tipo VARCHAR(30) DEFAULT 'comentario',
                fecha DATETIME DEFAULT NOW(),
                INDEX idx_formulario (formulario_id)
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS garantia_piezas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nombre VARCHAR(200) NOT NULL UNIQUE,
                activo TINYINT(1) DEFAULT 1,
                fecha_creacion DATETIME DEFAULT NOW()
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        conn.commit()
        # Sembrar piezas por defecto si la tabla está vacía
        cursor.execute("SELECT COUNT(*) as cnt FROM garantia_piezas")
        if cursor.fetchone()['cnt'] == 0:
            piezas_default = [
                'N/A', 'ASIENTO', 'BATERIA', 'CUADRO', 'DROPPER', 'DROPPER POST',
                'FRENOS', 'GOOGLES', 'GUANTES', 'HANGER', 'LLANTA', 'MANDO E-BIKE',
                'MAUBRIO', 'POTENCIA', 'RINES', 'SUSPENSION', 'TRANSMISION',
                'TWINLOCK', 'UNIDAD MOTRIZ', 'ZAPATOS',
            ]
            for p in piezas_default:
                cursor.execute("INSERT IGNORE INTO garantia_piezas (nombre) VALUES (%s)", (p,))
            conn.commit()

        # Columnas de seguimiento ampliado — se agregan si no existen
        for col_sql in [
            "ALTER TABLE garantia_formularios ADD COLUMN estatus_pieza VARCHAR(50) DEFAULT 'Sin pieza'",
            "ALTER TABLE garantia_formularios ADD COLUMN docs_validados TINYINT(1) DEFAULT 0",
            "ALTER TABLE garantia_formularios ADD COLUMN serie_validada TINYINT(1) DEFAULT 0",
            "ALTER TABLE garantia_formularios ADD COLUMN validacion_docs_json LONGTEXT DEFAULT NULL",
            "ALTER TABLE garantia_formularios ADD COLUMN pieza_reemplazo VARCHAR(100) DEFAULT NULL",
            "ALTER TABLE garantia_formularios ADD COLUMN fecha_estatus DATE DEFAULT NULL",
            "ALTER TABLE garantia_formularios ADD COLUMN fecha_pieza DATE DEFAULT NULL",
        ]:
            try:
                cursor.execute(col_sql)
                conn.commit()
            except Exception:
                conn.rollback()  # columna ya existe, ignorar
        return jsonify({"ok": True, "mensaje": "Tablas creadas correctamente"})
    except Exception as e:
        logging.exception("Error al inicializar tablas de garantias: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Form submissions ──────────────────────────────────────────────────────────

@garantias_bp.route("/formulario/enviar", methods=["POST"])
def enviar_formulario():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}

        # Determinar el email del ticket (a quién pertenece)
        email = datos.get('email', '')
        auth_header = request.headers.get('Authorization', '')
        if auth_header:
            raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
            if raw_token:
                payload = verificar_token(raw_token)
                if payload and payload.get('id'):
                    cursor_u = conn.cursor(dictionary=True)
                    cursor_u.execute("SELECT correo, rol_id FROM usuarios WHERE id = %s", (payload['id'],))
                    user = cursor_u.fetchone()
                    if user and user.get('correo'):
                        email = user['correo']
                        # Si el administrador asigna explícitamente a otro usuario, usar ese email
                        if user.get('rol_id') == 1 and datos.get('email_asignado'):
                            email = datos['email_asignado']

        cursor = conn.cursor()
        fecha_ingreso = datos.get('fecha_ingreso') or None

        # Insertar sin folio para obtener el ID auto-increment
        if fecha_ingreso:
            cursor.execute("""
                INSERT INTO garantia_formularios
                    (email, distribuidor, contacto, puesto, marca, datos, fecha_creacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                email,
                datos.get('distribuidor', ''),
                datos.get('contacto', ''),
                datos.get('puesto', ''),
                datos.get('marca', ''),
                json.dumps(datos, ensure_ascii=False),
                fecha_ingreso,
            ))
        else:
            cursor.execute("""
                INSERT INTO garantia_formularios
                    (email, distribuidor, contacto, puesto, marca, datos)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                email,
                datos.get('distribuidor', ''),
                datos.get('contacto', ''),
                datos.get('puesto', ''),
                datos.get('marca', ''),
                json.dumps(datos, ensure_ascii=False),
            ))
        new_id = cursor.lastrowid
        folio = f"GAR-{new_id:04d}"
        cursor.execute("UPDATE garantia_formularios SET folio = %s WHERE id = %s", (folio, new_id))
        conn.commit()
        return jsonify({"ok": True, "folio": folio, "id": new_id})
    except Exception as e:
        logging.exception("Error al enviar formulario de garantias: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/lista", methods=["GET"])
def lista_formularios():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, folio, email, distribuidor, contacto, puesto, marca,
                   estatus, estatus_pieza, pieza_reemplazo,
                   docs_validados, serie_validada,
                   validacion_docs_json, fecha_creacion,
                   fecha_estatus, fecha_pieza
            FROM garantia_formularios
            ORDER BY fecha_creacion DESC
        """)
        rows = cursor.fetchall()
        for r in rows:
            if r.get('fecha_creacion'):
                r['fecha_creacion'] = r['fecha_creacion'].strftime('%d/%m/%Y %H:%M')
            if r.get('fecha_estatus'):
                r['fecha_estatus'] = r['fecha_estatus'].strftime('%Y-%m-%d')
            if r.get('fecha_pieza'):
                r['fecha_pieza'] = r['fecha_pieza'].strftime('%Y-%m-%d')
        return jsonify(rows)
    except Exception as e:
        logging.exception("Error al listar formularios de garantias: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/actualizar-dato", methods=["PUT"])
def actualizar_dato_usuario(form_id):
    """El usuario dueño del ticket actualiza un campo rechazado y resetea su validación."""
    auth_header = request.headers.get('Authorization', '')
    raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
    if not raw_token:
        return jsonify({"error": "No autorizado"}), 401
    payload = verificar_token(raw_token)
    if not payload or not payload.get('id'):
        return jsonify({"error": "Token inválido"}), 401

    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT correo FROM usuarios WHERE id = %s", (payload['id'],))
        user = cursor.fetchone()
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        cursor.execute(
            "SELECT datos, validacion_docs_json FROM garantia_formularios WHERE id = %s AND email = %s",
            (form_id, user['correo'])
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Ticket no encontrado o no autorizado"}), 404

        body = request.get_json(force=True) or {}
        campo = body.get('campo', '')
        valor = body.get('valor')
        if not campo or valor is None:
            return jsonify({"error": "Campos requeridos"}), 400

        datos = {}
        if row.get('datos'):
            try:
                datos = json.loads(row['datos'])
            except Exception:
                datos = {}
        datos[campo] = valor

        val_json = {}
        if row.get('validacion_docs_json'):
            try:
                val_json = json.loads(row['validacion_docs_json'])
            except Exception:
                val_json = {}

        if campo == 'bici_serie':
            val_json.pop('numero_serie', None)
            nombre_legible = 'Número de Serie'
        else:
            val_json.pop(campo, None)
            nombre_legible = campo

        cursor.execute(
            "UPDATE garantia_formularios SET datos = %s, validacion_docs_json = %s WHERE id = %s",
            (json.dumps(datos, ensure_ascii=False), json.dumps(val_json, ensure_ascii=False), form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s,%s,%s,%s)",
            (form_id, 'Cliente', f'"{nombre_legible}" actualizado por el cliente', 'validacion')
        )
        conn.commit()
        return jsonify({"ok": True, "validacion_docs_json": val_json})
    except Exception as e:
        logging.exception("Error al actualizar dato usuario: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/mis-tickets", methods=["GET"])
def mis_tickets():
    """Devuelve los tickets del usuario autenticado (filtra por su correo registrado)."""
    auth_header = request.headers.get('Authorization', '')
    raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
    if not raw_token:
        return jsonify({"error": "No autorizado"}), 401
    payload = verificar_token(raw_token)
    if not payload:
        return jsonify({"error": "Token inválido"}), 401
    user_id = payload.get('id')
    if not user_id:
        return jsonify({"error": "Token inválido"}), 401

    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT correo FROM usuarios WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404
        email = user['correo']

        cursor.execute("""
            SELECT id, folio, email, distribuidor, contacto, puesto, marca,
                   estatus, estatus_pieza, fecha_creacion
            FROM garantia_formularios
            WHERE email = %s
            ORDER BY fecha_creacion DESC
        """, (email,))
        rows = cursor.fetchall()
        for r in rows:
            if r.get('fecha_creacion'):
                r['fecha_creacion'] = r['fecha_creacion'].strftime('%d/%m/%Y %H:%M')
        return jsonify(rows)
    except Exception as e:
        logging.exception("Error en mis-tickets: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>", methods=["DELETE"])
def eliminar_formulario(form_id):
    """Elimina un ticket y renumera los folios consecutivamente. Solo admins."""
    auth_header = request.headers.get('Authorization', '')
    raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
    if not raw_token:
        return jsonify({"error": "No autorizado"}), 401
    payload = verificar_token(raw_token)
    if not payload or not payload.get('id'):
        return jsonify({"error": "Token inválido"}), 401

    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor_u = conn.cursor(dictionary=True)
        cursor_u.execute("SELECT rol_id FROM usuarios WHERE id = %s", (payload['id'],))
        user = cursor_u.fetchone()
        if not user or user.get('rol_id') != 1:
            return jsonify({"error": "Solo administradores pueden eliminar tickets"}), 403

        cursor = conn.cursor()
        cursor.execute("DELETE FROM garantia_comentarios WHERE formulario_id = %s", (form_id,))
        cursor.execute("DELETE FROM garantia_formularios WHERE id = %s", (form_id,))
        if cursor.rowcount == 0:
            conn.rollback()
            return jsonify({"error": "No encontrado"}), 404

        # Renumerar consecutivamente por orden de id ascendente
        cursor.execute("SELECT id FROM garantia_formularios ORDER BY id ASC")
        ids = [row[0] for row in cursor.fetchall()]
        for i, record_id in enumerate(ids, start=1):
            cursor.execute(
                "UPDATE garantia_formularios SET folio = %s WHERE id = %s",
                (f"GAR-{i:04d}", record_id)
            )

        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        logging.exception("Error al eliminar formulario: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>", methods=["GET"])
def obtener_formulario(form_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM garantia_formularios WHERE id = %s", (form_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "No encontrado"}), 404
        if row.get('fecha_creacion'):
            row['fecha_creacion'] = row['fecha_creacion'].strftime('%d/%m/%Y %H:%M')
        if row.get('fecha_actualizacion'):
            row['fecha_actualizacion'] = row['fecha_actualizacion'].strftime('%d/%m/%Y %H:%M')
        if row.get('fecha_estatus'):
            row['fecha_estatus'] = row['fecha_estatus'].strftime('%Y-%m-%d')
        if row.get('fecha_pieza'):
            row['fecha_pieza'] = row['fecha_pieza'].strftime('%Y-%m-%d')
        if row.get('datos') and isinstance(row['datos'], str):
            row['datos'] = json.loads(row['datos'])
        if row.get('validacion_docs_json') and isinstance(row['validacion_docs_json'], str):
            row['validacion_docs_json'] = json.loads(row['validacion_docs_json'])
        return jsonify(row)
    except Exception as e:
        logging.exception("Error al obtener formulario: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/estatus", methods=["PUT"])
def actualizar_estatus(form_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        nuevo_estatus = datos.get('estatus', 'Enviado')
        fecha = datos.get('fecha') or date.today().isoformat()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE garantia_formularios SET estatus = %s, fecha_estatus = %s WHERE id = %s",
            (nuevo_estatus, fecha, form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s,%s,%s,%s)",
            (form_id, 'Sistema', f'Estatus actualizado a "{nuevo_estatus}" (fecha: {fecha})', 'estatus')
        )
        conn.commit()
        return jsonify({"ok": True, "fecha_estatus": fecha})
    except Exception as e:
        logging.exception("Error al actualizar estatus: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/fecha-estatus", methods=["PUT"])
def actualizar_fecha_estatus(form_id):
    """Edita solo la fecha del estatus actual sin cambiar el estatus."""
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        fecha = datos.get('fecha') or date.today().isoformat()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE garantia_formularios SET fecha_estatus = %s WHERE id = %s",
            (fecha, form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s,%s,%s,%s)",
            (form_id, 'Sistema', f'Fecha de estatus corregida a {fecha}', 'estatus')
        )
        conn.commit()
        return jsonify({"ok": True, "fecha_estatus": fecha})
    except Exception as e:
        logging.exception("Error al actualizar fecha_estatus: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/pieza-reemplazo", methods=["PUT"])
def actualizar_pieza_reemplazo(form_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        pieza = datos.get('pieza_reemplazo', '')
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE garantia_formularios SET pieza_reemplazo = %s WHERE id = %s",
            (pieza, form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s,%s,%s,%s)",
            (form_id, 'Sistema', f'Pieza de reemplazo asignada: "{pieza}"', 'pieza')
        )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception("Error al actualizar pieza_reemplazo: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/pieza", methods=["PUT"])
def actualizar_pieza(form_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        nuevo_estatus = datos.get('estatus_pieza', 'Sin pieza')
        fecha = datos.get('fecha') or date.today().isoformat()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE garantia_formularios SET estatus_pieza = %s, fecha_pieza = %s WHERE id = %s",
            (nuevo_estatus, fecha, form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s, %s, %s, %s)",
            (form_id, 'Sistema', f'Estado de pieza actualizado a "{nuevo_estatus}" (fecha: {fecha})', 'pieza')
        )
        conn.commit()
        return jsonify({"ok": True, "fecha_pieza": fecha})
    except Exception as e:
        logging.exception("Error al actualizar pieza: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/fecha-pieza", methods=["PUT"])
def actualizar_fecha_pieza(form_id):
    """Edita solo la fecha del estatus de pieza sin cambiar el estatus."""
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        fecha = datos.get('fecha') or date.today().isoformat()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE garantia_formularios SET fecha_pieza = %s WHERE id = %s",
            (fecha, form_id)
        )
        cursor.execute(
            "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s, %s, %s, %s)",
            (form_id, 'Sistema', f'Fecha de pieza corregida a {fecha}', 'pieza')
        )
        conn.commit()
        return jsonify({"ok": True, "fecha_pieza": fecha})
    except Exception as e:
        logging.exception("Error al actualizar fecha_pieza: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/validacion-doc", methods=["PUT"])
def actualizar_validacion_doc(form_id):
    """Valida o rechaza un documento individual. body: {campo, estado, nombre_legible}"""
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        campo          = datos.get('campo', '')
        estado         = datos.get('estado')          # 'valido' | 'rechazado' | None
        nombre_legible = datos.get('nombre_legible', campo)
        fecha_val      = datos.get('fecha_validacion') or None  # YYYY-MM-DD opcional

        if not campo:
            return jsonify({"error": "Campo requerido"}), 400

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT validacion_docs_json FROM garantia_formularios WHERE id = %s",
            (form_id,)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "No encontrado"}), 404

        val_json = {}
        if row.get('validacion_docs_json'):
            try:
                val_json = json.loads(row['validacion_docs_json'])
            except Exception:
                val_json = {}

        if estado is None:
            val_json.pop(campo, None)
        else:
            val_json[campo] = estado

        cursor.execute(
            "UPDATE garantia_formularios SET validacion_docs_json = %s WHERE id = %s",
            (json.dumps(val_json, ensure_ascii=False), form_id)
        )

        accion = 'validado' if estado == 'valido' else ('rechazado' if estado == 'rechazado' else 'restablecido')
        texto  = f'Documento "{nombre_legible}" {accion}'
        if fecha_val:
            cursor.execute(
                "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo, fecha) VALUES (%s, %s, %s, %s, %s)",
                (form_id, 'Sistema', texto, 'validacion', fecha_val)
            )
        else:
            cursor.execute(
                "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s, %s, %s, %s)",
                (form_id, 'Sistema', texto, 'validacion')
            )
        conn.commit()
        return jsonify({"ok": True, "validacion_docs_json": val_json})
    except Exception as e:
        logging.exception("Error al validar documento: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/formulario/<int:form_id>/validacion", methods=["PUT"])
def actualizar_validacion(form_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        docs_validados = 1 if datos.get('docs_validados') else 0
        serie_validada = 1 if datos.get('serie_validada') else 0

        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT docs_validados, serie_validada FROM garantia_formularios WHERE id = %s",
            (form_id,)
        )
        prev = cursor.fetchone() or {}

        cursor.execute(
            "UPDATE garantia_formularios SET docs_validados = %s, serie_validada = %s WHERE id = %s",
            (docs_validados, serie_validada, form_id)
        )

        # Registrar en historial sólo los cambios reales
        if int(prev.get('docs_validados', 0)) != docs_validados:
            estado = 'validados' if docs_validados else 'marcados como inválidos'
            cursor.execute(
                "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s, %s, %s, %s)",
                (form_id, 'Sistema', f'Documentos {estado}', 'validacion')
            )
        if int(prev.get('serie_validada', 0)) != serie_validada:
            estado = 'validado' if serie_validada else 'marcado como inválido'
            cursor.execute(
                "INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo) VALUES (%s, %s, %s, %s)",
                (form_id, 'Sistema', f'Número de serie {estado}', 'validacion')
            )

        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception("Error al actualizar validacion: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Form structure (editor) ───────────────────────────────────────────────────

@garantias_bp.route("/estructura", methods=["GET"])
def obtener_estructura():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM garantia_estructura ORDER BY id DESC LIMIT 1")
        row = cursor.fetchone()
        if not row:
            return jsonify({"estructura": None, "version": 0})
        if row.get('estructura') and isinstance(row['estructura'], str):
            row['estructura'] = json.loads(row['estructura'])
        if row.get('fecha_creacion'):
            row['fecha_creacion'] = row['fecha_creacion'].strftime('%d/%m/%Y %H:%M')
        return jsonify(row)
    except Exception as e:
        logging.exception("Error al obtener estructura: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/estructura", methods=["POST"])
def guardar_estructura():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT MAX(version) as mv FROM garantia_estructura")
        row = cursor.fetchone()
        nueva_version = (row['mv'] or 0) + 1

        cursor.execute(
            "INSERT INTO garantia_estructura (version, estructura) VALUES (%s, %s)",
            (nueva_version, json.dumps(datos.get('estructura'), ensure_ascii=False))
        )
        conn.commit()
        return jsonify({"ok": True, "version": nueva_version})
    except Exception as e:
        logging.exception("Error al guardar estructura: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Stats (hub) ──────────────────────────────────────────────────────────────

@garantias_bp.route("/stats", methods=["GET"])
def get_stats():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) as total FROM garantia_formularios")
        total = cursor.fetchone()['total']
        cursor.execute("""
            SELECT COUNT(*) as abiertos FROM garantia_formularios
            WHERE estatus IN ('Enviado', 'En revisión')
        """)
        abiertos = cursor.fetchone()['abiertos']
        cursor.execute("""
            SELECT COUNT(*) as este_mes FROM garantia_formularios
            WHERE MONTH(fecha_creacion) = MONTH(NOW())
              AND YEAR(fecha_creacion) = YEAR(NOW())
        """)
        este_mes = cursor.fetchone()['este_mes']
        cursor.execute("""
            SELECT COUNT(*) as cerrados FROM garantia_formularios
            WHERE estatus IN ('Cerrado', 'Aprobado', 'Rechazado')
        """)
        cerrados = cursor.fetchone()['cerrados']
        return jsonify({"total": total, "abiertos": abiertos, "este_mes": este_mes, "cerrados": cerrados})
    except Exception as e:
        logging.exception("Error en stats: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Comentarios de ticket ─────────────────────────────────────────────────────

@garantias_bp.route("/ticket/<int:formulario_id>/comentarios", methods=["GET"])
def get_comentarios(formulario_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        es_admin = False
        auth_header = request.headers.get('Authorization', '')
        if auth_header:
            raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
            if raw_token:
                payload = verificar_token(raw_token)
                if payload and payload.get('id'):
                    cursor_u = conn.cursor(dictionary=True)
                    cursor_u.execute("SELECT rol_id FROM usuarios WHERE id = %s", (payload['id'],))
                    user = cursor_u.fetchone()
                    es_admin = bool(user and user.get('rol_id') == 1)

        cursor = conn.cursor(dictionary=True)
        if es_admin:
            cursor.execute("""
                SELECT * FROM garantia_comentarios
                WHERE formulario_id = %s ORDER BY fecha ASC
            """, (formulario_id,))
        else:
            cursor.execute("""
                SELECT * FROM garantia_comentarios
                WHERE formulario_id = %s AND tipo != 'nota_interna'
                ORDER BY fecha ASC
            """, (formulario_id,))
        rows = cursor.fetchall()
        for r in rows:
            if r.get('fecha'):
                r['fecha'] = r['fecha'].strftime('%d/%m/%Y %H:%M')
        return jsonify(rows)
    except Exception as e:
        logging.exception("Error al obtener comentarios: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/ticket/<int:formulario_id>/comentarios", methods=["POST"])
def add_comentario(formulario_id):
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        texto = (datos.get('texto') or '').strip()
        if not texto:
            return jsonify({"error": "El comentario no puede estar vacio"}), 400
        tipo = datos.get('tipo', 'comentario')

        if tipo == 'nota_interna':
            # Extraer nombre real del JWT — requerido para notas internas
            auth_header = request.headers.get('Authorization', '')
            raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
            if not raw_token:
                return jsonify({"error": "No autorizado"}), 401
            tok_payload = verificar_token(raw_token)
            if not tok_payload or not tok_payload.get('id'):
                return jsonify({"error": "Token inválido"}), 401
            cursor_u = conn.cursor(dictionary=True)
            cursor_u.execute("SELECT nombre, rol_id FROM usuarios WHERE id = %s", (tok_payload['id'],))
            user = cursor_u.fetchone()
            if not user or user.get('rol_id') != 1:
                return jsonify({"error": "No autorizado"}), 403
            autor = user.get('nombre') or 'Administrador'
        else:
            autor = (datos.get('autor') or 'Administrador').strip()

        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo)
            VALUES (%s, %s, %s, %s)
        """, (formulario_id, autor, texto, tipo))
        conn.commit()
        return jsonify({"ok": True, "id": cursor.lastrowid})
    except Exception as e:
        logging.exception("Error al agregar comentario: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Latencias por ticket ─────────────────────────────────────────────────────

@garantias_bp.route("/latencias", methods=["GET"])
def get_latencias():
    """Devuelve latencia de atención y de cierre por ticket individual."""
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                f.id, f.folio, f.estatus, f.distribuidor,
                CASE WHEN f.estatus IN ('Cerrado','Rechazado')
                     THEN DATEDIFF(
                         COALESCE(f.fecha_estatus, DATE(f.fecha_actualizacion)),
                         DATE(f.fecha_creacion)
                     )
                     ELSE NULL END AS lat_cierre,
                MIN(CASE WHEN c.tipo = 'validacion' AND DATEDIFF(c.fecha, f.fecha_creacion) >= 0
                         THEN DATEDIFF(c.fecha, f.fecha_creacion)
                         ELSE NULL END) AS lat_atencion
            FROM garantia_formularios f
            LEFT JOIN garantia_comentarios c ON c.formulario_id = f.id
            GROUP BY f.id, f.folio, f.estatus, f.distribuidor,
                     f.fecha_creacion, f.fecha_actualizacion, f.fecha_estatus
            ORDER BY f.fecha_creacion DESC
        """)
        rows = cursor.fetchall()
        return jsonify(rows)
    except Exception as e:
        logging.exception("Error en /garantias/latencias: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Catálogo de piezas de reemplazo ──────────────────────────────────────────

@garantias_bp.route("/piezas", methods=["GET"])
def listar_piezas():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS garantia_piezas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nombre VARCHAR(200) NOT NULL UNIQUE,
                activo TINYINT(1) DEFAULT 1,
                fecha_creacion DATETIME DEFAULT NOW()
            ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """)
        cursor.execute("SELECT COUNT(*) as cnt FROM garantia_piezas")
        if cursor.fetchone()['cnt'] == 0:
            piezas_default = [
                'N/A', 'ASIENTO', 'BATERIA', 'CUADRO', 'DROPPER', 'DROPPER POST',
                'FRENOS', 'GOOGLES', 'GUANTES', 'HANGER', 'LLANTA', 'MANDO E-BIKE',
                'MAUBRIO', 'POTENCIA', 'RINES', 'SUSPENSION', 'TRANSMISION',
                'TWINLOCK', 'UNIDAD MOTRIZ', 'ZAPATOS',
            ]
            for p in piezas_default:
                cursor.execute("INSERT IGNORE INTO garantia_piezas (nombre) VALUES (%s)", (p,))
        conn.commit()
        cursor.execute("""
            SELECT nombre FROM garantia_piezas
            WHERE activo = 1
            ORDER BY nombre = 'N/A' DESC, nombre ASC
        """)
        return jsonify([r['nombre'] for r in cursor.fetchall()])
    except Exception as e:
        logging.exception("Error al listar piezas: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@garantias_bp.route("/piezas", methods=["POST"])
def agregar_pieza():
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexion a BD"}), 500
    try:
        datos = request.get_json(force=True) or {}
        nombre = (datos.get('nombre') or '').strip().upper()
        if not nombre:
            return jsonify({"error": "El nombre de la pieza es requerido"}), 400
        cursor = conn.cursor()
        cursor.execute("INSERT IGNORE INTO garantia_piezas (nombre) VALUES (%s)", (nombre,))
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({"ok": True, "mensaje": "La pieza ya existía", "nombre": nombre})
        return jsonify({"ok": True, "nombre": nombre})
    except Exception as e:
        logging.exception("Error al agregar pieza: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── Importación masiva histórica ──────────────────────────────────────────────

COLUMNAS_IMPORTACION = [
    ("folio_anterior",    "Folio anterior (referencia)",   "GAR-001"),
    ("distribuidor",      "Distribuidor *",                "Elite Bike GDL"),
    ("contacto",          "Contacto *",                    "Juan López"),
    ("puesto",            "Puesto",                        "Gerente"),
    ("marca",             "Marca *",                       "Scott"),
    ("modelo",            "Modelo del producto",           "Scale 960"),
    ("descripcion_dano",  "Descripción del daño *",        "Fisura en tubo inferior del cuadro"),
    ("estatus",           "Estatus *",                     "Cerrado"),
    ("estatus_pieza",     "Estatus pieza",                 "Enviada al cliente"),
    ("pieza_reemplazo",   "Pieza de reemplazo",            "CUADRO"),
    ("fecha_registro",    "Fecha de registro *",           "2024-03-15"),
    ("fecha_atencion",    "Fecha de atención (validación docs)", "2024-03-22"),
    ("fecha_cierre",      "Fecha de cierre",               "2024-04-02"),
    ("asignado_a",        "Asignado a (nombre exacto del usuario)", "Victor Alejandro Garnier Morga"),
]

ESTATUSES_VALIDOS      = ['Enviado', 'En revisión', 'Aprobado', 'Rechazado', 'Cerrado']
ESTATUSES_PIEZA_VALIDOS = ['Sin pieza', 'Solicitada', 'En tránsito', 'En almacén', 'Enviada al cliente', 'Rechazada']
MARCAS_VALIDAS         = ['Scott', 'Megamo', 'Syncros', 'Vittoria', 'Bosch', 'Otra']


def _requiere_admin(request):
    """Retorna (payload, None) si es admin, o (None, response_error) si no."""
    auth_header = request.headers.get('Authorization', '')
    raw_token = auth_header.split(' ')[1] if ' ' in auth_header else None
    if not raw_token:
        return None, (jsonify({"error": "No autorizado"}), 401)
    payload = verificar_token(raw_token)
    if not payload or not payload.get('id'):
        return None, (jsonify({"error": "Token inválido"}), 401)
    conn = obtener_conexion()
    if not conn:
        return None, (jsonify({"error": "Sin conexión a BD"}), 500)
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT rol_id FROM usuarios WHERE id = %s", (payload['id'],))
        user = cursor.fetchone()
    finally:
        conn.close()
    if not user or user.get('rol_id') != 1:
        return None, (jsonify({"error": "Solo administradores"}), 403)
    return payload, None


@garantias_bp.route("/importar/plantilla", methods=["GET"])
def descargar_plantilla_importacion():
    """Genera y devuelve el Excel plantilla para importación masiva de garantías."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    _, err = _requiere_admin(request)
    if err:
        return err

    wb = Workbook()

    # ── Hoja 1: Importar ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Importar"

    hdr_fill   = PatternFill("solid", fgColor="1A1A2E")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    req_fill   = PatternFill("solid", fgColor="C8102E")
    ex_fill    = PatternFill("solid", fgColor="2A2A4E")
    ex_font    = Font(color="AAAACC", italic=True, size=10)
    thin       = Side(style='thin', color="444466")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",  vertical="center", wrap_text=True)

    # Fila 1: instrucción general
    ws.merge_cells("A1:N1")
    ws["A1"] = (
        "PLANTILLA DE IMPORTACIÓN MASIVA — GARANTÍAS ELITE BIKE  |  "
        "Campos con * son obligatorios  |  Fechas en formato YYYY-MM-DD (ej. 2024-03-15)  |  "
        "Consulta la hoja 'Usuarios' para los nombres exactos asignables"
    )
    ws["A1"].fill      = PatternFill("solid", fgColor="0D0D1A")
    ws["A1"].font      = Font(bold=True, color="FFD700", size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Fila 2: encabezados
    for col_i, (_, label, _) in enumerate(COLUMNAS_IMPORTACION, start=1):
        cell         = ws.cell(row=2, column=col_i, value=label)
        cell.fill    = req_fill if label.endswith("*") else hdr_fill
        cell.font    = hdr_font
        cell.alignment = center
        cell.border  = border
    ws.row_dimensions[2].height = 36

    # Fila 3: fila de ejemplo
    for col_i, (_, _, ejemplo) in enumerate(COLUMNAS_IMPORTACION, start=1):
        cell           = ws.cell(row=3, column=col_i, value=ejemplo)
        cell.fill      = ex_fill
        cell.font      = ex_font
        cell.alignment = left
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # Anchos de columna
    anchos = [15, 25, 22, 18, 14, 20, 40, 15, 22, 20, 16, 16, 16, 30]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Validaciones desplegables en columnas de estatus y marca (filas 4-300)
    def add_dv(formula, col_letter):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.sqref = f"{col_letter}4:{col_letter}300"
        ws.add_data_validation(dv)

    col_estatus      = get_column_letter(8)   # estatus
    col_est_pieza    = get_column_letter(9)   # estatus_pieza
    col_marca        = get_column_letter(5)   # marca

    add_dv('"' + ','.join(ESTATUSES_VALIDOS) + '"',       col_estatus)
    add_dv('"' + ','.join(ESTATUSES_PIEZA_VALIDOS) + '"', col_est_pieza)
    add_dv('"' + ','.join(MARCAS_VALIDAS) + '"',          col_marca)

    # ── Hoja 2: Usuarios ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Usuarios disponibles")
    ws2["A1"] = "Nombre exacto del usuario"
    ws2["B1"] = "Correo"
    ws2["A1"].font = Font(bold=True, color="FFFFFF")
    ws2["B1"].font = Font(bold=True, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1A1A2E")
    ws2["B1"].fill = PatternFill("solid", fgColor="1A1A2E")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 35

    conn = obtener_conexion()
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            cur.execute("SELECT nombre, correo FROM usuarios WHERE activo = 1 ORDER BY nombre")
            for i, u in enumerate(cur.fetchall(), start=2):
                ws2.cell(row=i, column=1, value=u.get('nombre') or '')
                ws2.cell(row=i, column=2, value=u.get('correo') or '')
        finally:
            conn.close()

    # ── Hoja 3: Valores válidos ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Valores válidos")
    secciones = [
        ("Estatus del ticket",    ESTATUSES_VALIDOS),
        ("Estatus de la pieza",   ESTATUSES_PIEZA_VALIDOS),
        ("Marcas válidas",        MARCAS_VALIDAS),
    ]
    col_offset = 1
    for titulo, valores in secciones:
        ws3.cell(row=1, column=col_offset, value=titulo).font = Font(bold=True)
        for i, v in enumerate(valores, start=2):
            ws3.cell(row=i, column=col_offset, value=v)
        ws3.column_dimensions[get_column_letter(col_offset)].width = 25
        col_offset += 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_importacion_garantias.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@garantias_bp.route("/importar", methods=["POST"])
def importar_garantias():
    """Recibe el Excel completado y hace la inserción masiva."""
    from openpyxl import load_workbook

    _, err = _requiere_admin(request)
    if err:
        return err

    if 'archivo' not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    file = request.files['archivo']
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        return jsonify({"error": "El archivo debe ser .xlsx"}), 400

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el Excel: {e}"}), 400

    ws = wb["Importar"] if "Importar" in wb.sheetnames else wb.active

    # Cargar usuarios del sistema para mapeo nombre → correo
    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexión a BD"}), 500

    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, nombre, correo FROM usuarios WHERE activo = 1")
        usuarios_db = cur.fetchall()
    finally:
        conn.close()

    # Índice nombre_normalizado → correo
    def _norm(s):
        return str(s or '').strip().lower()

    usuarios_idx = {_norm(u['nombre']): u['correo'] for u in usuarios_db}

    insertados = 0
    errores    = []
    filas_datos = list(ws.iter_rows(min_row=4, values_only=True))  # fila 1=instrucción, 2=header, 3=ejemplo

    conn = obtener_conexion()
    if not conn:
        return jsonify({"error": "Sin conexión a BD"}), 500

    try:
        cursor = conn.cursor()

        for fila_num, row in enumerate(filas_datos, start=4):
            # Ignorar filas completamente vacías
            if all(c is None or str(c).strip() == '' for c in row):
                continue

            def cel(idx):
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ''

            distribuidor    = cel(1)
            contacto        = cel(2)
            puesto          = cel(3)
            marca           = cel(4)
            modelo          = cel(5)
            descripcion     = cel(6)
            estatus         = cel(7) or 'Enviado'
            estatus_pieza   = cel(8) or 'Sin pieza'
            pieza_reemplazo = cel(9)
            fecha_registro  = cel(10)
            fecha_atencion  = cel(11)
            fecha_cierre    = cel(12)
            asignado_a      = cel(13)

            # Validaciones obligatorias
            fila_err = []
            if not distribuidor:  fila_err.append("Falta distribuidor")
            if not contacto:      fila_err.append("Falta contacto")
            if not marca:         fila_err.append("Falta marca")
            if not descripcion:   fila_err.append("Falta descripción del daño")
            if not fecha_registro: fila_err.append("Falta fecha de registro")
            if estatus not in ESTATUSES_VALIDOS:
                fila_err.append(f"Estatus inválido: '{estatus}'")

            # Parsear fecha_registro
            fecha_creacion_dt = None
            if fecha_registro:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        fecha_creacion_dt = datetime.strptime(fecha_registro, fmt)
                        break
                    except ValueError:
                        continue
                if not fecha_creacion_dt:
                    fila_err.append(f"Fecha de registro inválida: '{fecha_registro}' (use YYYY-MM-DD)")

            if fila_err:
                errores.append({"fila": fila_num, "errores": fila_err})
                continue

            # Resolver email del asignado
            email_ticket = ''
            if asignado_a:
                email_ticket = usuarios_idx.get(_norm(asignado_a), '')
                if not email_ticket:
                    errores.append({
                        "fila": fila_num,
                        "errores": [f"Usuario '{asignado_a}' no encontrado — ticket importado sin asignar"]
                    })

            # Construir JSON de datos del formulario
            datos_json = json.dumps({
                "distribuidor":     distribuidor,
                "contacto":         contacto,
                "puesto":           puesto,
                "marca":            marca,
                "modelo":           modelo,
                "descripcion_dano": descripcion,
                "importado":        True,
            }, ensure_ascii=False)

            # Insertar registro principal
            cursor.execute("""
                INSERT INTO garantia_formularios
                    (email, distribuidor, contacto, puesto, marca,
                     datos, estatus, estatus_pieza, pieza_reemplazo, fecha_creacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                email_ticket, distribuidor, contacto, puesto, marca,
                datos_json, estatus,
                estatus_pieza if estatus_pieza in ESTATUSES_PIEZA_VALIDOS else 'Sin pieza',
                pieza_reemplazo or None,
                fecha_creacion_dt,
            ))
            new_id = cursor.lastrowid
            folio  = f"GAR-{new_id:04d}"
            cursor.execute("UPDATE garantia_formularios SET folio = %s WHERE id = %s", (folio, new_id))

            # Fecha de cierre → fecha_actualizacion si estatus es terminal
            if fecha_cierre and estatus in ('Cerrado', 'Rechazado'):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        dt_cierre = datetime.strptime(fecha_cierre, fmt)
                        cursor.execute(
                            "UPDATE garantia_formularios SET fecha_actualizacion = %s WHERE id = %s",
                            (dt_cierre, new_id)
                        )
                        break
                    except ValueError:
                        continue

            # Fecha de atención → comentario tipo 'validacion' para la latencia
            if fecha_atencion:
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                    try:
                        dt_atencion = datetime.strptime(fecha_atencion, fmt)
                        cursor.execute("""
                            INSERT INTO garantia_comentarios (formulario_id, autor, texto, tipo, fecha)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (new_id, 'Sistema', 'Documentos validados (importación histórica)', 'validacion', dt_atencion))
                        break
                    except ValueError:
                        continue

            insertados += 1

        conn.commit()
        return jsonify({
            "ok":        True,
            "insertados": insertados,
            "errores":   errores,
            "mensaje":   f"{insertados} tickets importados correctamente" + (f", {len(errores)} con advertencias" if errores else ""),
        })

    except Exception as e:
        conn.rollback()
        logging.exception("Error en importación masiva de garantías: %s", e)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ── File uploads ──────────────────────────────────────────────────────────────

@garantias_bp.route("/archivo/subir", methods=["POST"])
def subir_archivo():
    if 'archivo' not in request.files:
        return jsonify({"error": "No se recibio archivo"}), 400

    file = request.files['archivo']

    if not file.filename:
        return jsonify({"error": "Nombre de archivo vacio"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Tipo de archivo no permitido"}), 400

    try:
        resultado = subir_archivo_s3(file)
        url = generar_url_firmada_s3(resultado["key"])

        return jsonify({
            "ok": True,
            "nombre": resultado["key"],
            "original": resultado["original"],
            "url": url,
            "storage": "s3"
        })

    except Exception as e:
        logging.exception("Error subiendo archivo de garantia a S3: %s", e)
        return jsonify({"error": "Error al subir archivo"}), 500


@garantias_bp.route("/archivo/<path:nombre>", methods=["GET"])
def descargar_archivo(nombre):
    try:
        key_s3 = nombre

        if not existe_archivo_s3(key_s3):
            return jsonify({"error": "Archivo no encontrado"}), 404

        url = generar_url_firmada_s3(key_s3)

        return redirect(url)

    except Exception as e:
        logging.exception("Error obteniendo archivo de garantia desde S3: %s", e)
        return jsonify({"error": "Error al obtener archivo"}), 500
