from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db_conexion import obtener_conexion


reportes_bp = Blueprint(
    "reportes_inventario",
    __name__,
    url_prefix="/api/inventario/reportes",
)


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================


def limpiar_texto(valor):
    if valor is None:
        return None

    if isinstance(valor, str):
        valor = valor.strip()
        return valor if valor else None

    return valor


def fecha_a_texto(valor):
    if not valor:
        return ""

    if isinstance(valor, datetime):
        return valor.isoformat(sep=" ")

    if isinstance(valor, date):
        return valor.isoformat()

    return str(valor)


def normalizar_fecha(valor, nombre_campo):
    valor = limpiar_texto(valor)

    if not valor:
        return None

    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError as error:
        raise ValueError(
            f"El campo {nombre_campo} debe tener el formato AAAA-MM-DD."
        ) from error


def nombre_completo(row):
    return " ".join(
        str(parte).strip()
        for parte in [
            row.get("nombre"),
            row.get("apellido_paterno"),
            row.get("apellido_materno"),
        ]
        if parte and str(parte).strip()
    )


def obtener_filtros():
    return {
        "empresa": limpiar_texto(request.args.get("empresa")),
        "departamento": limpiar_texto(request.args.get("departamento")),
        "estado": limpiar_texto(request.args.get("estado")),
        "fecha_inicio": normalizar_fecha(
            request.args.get("fechaInicio"),
            "fecha inicial",
        ),
        "fecha_fin": normalizar_fecha(
            request.args.get("fechaFin"),
            "fecha final",
        ),
        "busqueda": limpiar_texto(request.args.get("busqueda")),
    }


def construir_filtro_equipos(filtros, alias="e"):
    condiciones = []
    parametros = []

    if filtros.get("empresa") and filtros["empresa"] != "Todas":
        condiciones.append(f"{alias}.empresa = %s")
        parametros.append(filtros["empresa"])

    if filtros.get("departamento") and filtros["departamento"] != "Todos":
        condiciones.append(f"{alias}.departamento = %s")
        parametros.append(filtros["departamento"])

    if filtros.get("estado") and filtros["estado"] != "Todos":
        condiciones.append(f"{alias}.estado = %s")
        parametros.append(filtros["estado"])

    if filtros.get("busqueda"):
        termino = f"%{filtros['busqueda']}%"
        condiciones.append(
            f"""
            (
                {alias}.numero_inventario LIKE %s
                OR {alias}.descripcion LIKE %s
                OR {alias}.categoria LIKE %s
                OR {alias}.marca LIKE %s
                OR {alias}.modelo LIKE %s
                OR {alias}.numero_serie LIKE %s
                OR {alias}.responsable LIKE %s
                OR {alias}.ubicacion LIKE %s
            )
            """
        )
        parametros.extend([termino] * 8)

    clausula = ""
    if condiciones:
        clausula = " WHERE " + " AND ".join(condiciones)

    return clausula, parametros


def construir_filtro_fecha(
    filtros,
    columna,
    condiciones=None,
    parametros=None,
):
    condiciones = list(condiciones or [])
    parametros = list(parametros or [])

    if filtros.get("fecha_inicio"):
        condiciones.append(f"DATE({columna}) >= %s")
        parametros.append(filtros["fecha_inicio"])

    if filtros.get("fecha_fin"):
        condiciones.append(f"DATE({columna}) <= %s")
        parametros.append(filtros["fecha_fin"])

    return condiciones, parametros


def clausula_where(condiciones):
    return " WHERE " + " AND ".join(condiciones) if condiciones else ""


def obtener_columnas_tabla(cursor, tabla):
    cursor.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (tabla,),
    )
    return {fila["COLUMN_NAME"] for fila in cursor.fetchall()}


def obtener_columna_fecha_movimientos(cursor):
    columnas = obtener_columnas_tabla(cursor, "inventario_movimientos")

    for candidata in [
        "fecha_movimiento",
        "fecha_creacion",
        "fecha_registro",
        "created_at",
    ]:
        if candidata in columnas:
            return candidata

    return None


def obtener_columna_id_movimientos(cursor):
    columnas = obtener_columnas_tabla(cursor, "inventario_movimientos")

    for candidata in ["id", "movimiento_id"]:
        if candidata in columnas:
            return candidata

    return None


def tabla_existe(cursor, tabla):
    cursor.execute(
        """
        SELECT COUNT(*) AS total
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (tabla,),
    )
    return int((cursor.fetchone() or {}).get("total") or 0) > 0


def convertir_filas_json(filas):
    resultado = []

    for fila in filas:
        nueva = {}
        for clave, valor in fila.items():
            nueva[clave] = fecha_a_texto(valor) if isinstance(
                valor,
                (date, datetime),
            ) else valor
        resultado.append(nueva)

    return resultado


def ejecutar_conteo(cursor, consulta, parametros=()):
    cursor.execute(consulta, tuple(parametros))
    fila = cursor.fetchone() or {}
    return int(fila.get("total") or 0)


def preparar_hoja(hoja, titulo, encabezados, filas):
    hoja.title = titulo[:31]
    hoja.freeze_panes = "A2"

    relleno = PatternFill("solid", fgColor="FF5A1F")
    fuente = Font(color="FFFFFF", bold=True)

    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=encabezado)
        celda.fill = relleno
        celda.font = fuente
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila_indice, fila in enumerate(filas, start=2):
        for columna_indice, valor in enumerate(fila, start=1):
            hoja.cell(
                row=fila_indice,
                column=columna_indice,
                value=valor,
            )

    for indice, encabezado in enumerate(encabezados, start=1):
        ancho = len(str(encabezado)) + 2

        for fila in hoja.iter_rows(
            min_row=2,
            min_col=indice,
            max_col=indice,
        ):
            valor = fila[0].value
            if valor is not None:
                ancho = max(ancho, min(len(str(valor)) + 2, 55))

        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    hoja.auto_filter.ref = hoja.dimensions


# ============================================================
# GET: CATÁLOGOS DE FILTROS
# ============================================================


@reportes_bp.route("/catalogos", methods=["GET"])
def obtener_catalogos():
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    try:
        cursor.execute(
            """
            SELECT DISTINCT empresa
            FROM inventario_equipos
            WHERE empresa IS NOT NULL AND TRIM(empresa) <> ''
            ORDER BY empresa
            """
        )
        empresas = [fila["empresa"] for fila in cursor.fetchall()]

        cursor.execute(
            """
            SELECT DISTINCT departamento
            FROM inventario_equipos
            WHERE departamento IS NOT NULL AND TRIM(departamento) <> ''
            ORDER BY departamento
            """
        )
        departamentos = [
            fila["departamento"]
            for fila in cursor.fetchall()
        ]

        return jsonify(
            {
                "empresas": empresas,
                "departamentos": departamentos,
                "estadosEquipo": [
                    "Todos",
                    "Asignado",
                    "Disponible",
                    "Baja",
                ],
                "tiposReporte": [
                    "equipos",
                    "asignaciones",
                    "responsivas",
                    "auditorias",
                    "movimientos",
                ],
            }
        )

    except Exception as error:
        print("Error al obtener catálogos de reportes:", error)
        return jsonify(
            {
                "error": "No se pudieron cargar los catálogos.",
                "detalle": str(error),
            }
        ), 500

    finally:
        cursor.close()
        conexion.close()


# ============================================================
# GET: DASHBOARD GENERAL DE REPORTES
# ============================================================


@reportes_bp.route("/dashboard", methods=["GET"])
def obtener_dashboard():
    try:
        filtros = obtener_filtros()
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    try:
        filtro_equipos, parametros_equipos = construir_filtro_equipos(filtros)

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN e.estado = 'Asignado' THEN 1 ELSE 0 END) AS asignados,
                SUM(CASE WHEN e.estado = 'Disponible' THEN 1 ELSE 0 END) AS disponibles,
                SUM(CASE WHEN e.estado = 'Baja' THEN 1 ELSE 0 END) AS bajas,
                SUM(CASE WHEN e.responsiva_estado = 'Pendiente' THEN 1 ELSE 0 END) AS responsivas_pendientes,
                SUM(CASE WHEN e.responsiva_estado = 'Firmada' THEN 1 ELSE 0 END) AS responsivas_firmadas
            FROM inventario_equipos e
            """ + filtro_equipos,
            tuple(parametros_equipos),
        )
        resumen_equipos = cursor.fetchone() or {}

        condiciones_asignaciones = []
        parametros_asignaciones = []

        if filtros.get("empresa") and filtros["empresa"] != "Todas":
            condiciones_asignaciones.append("e.empresa = %s")
            parametros_asignaciones.append(filtros["empresa"])

        if filtros.get("departamento") and filtros["departamento"] != "Todos":
            condiciones_asignaciones.append("e.departamento = %s")
            parametros_asignaciones.append(filtros["departamento"])

        condiciones_asignaciones, parametros_asignaciones = construir_filtro_fecha(
            filtros,
            "a.fecha_asignacion",
            condiciones_asignaciones,
            parametros_asignaciones,
        )

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN a.estado = 'Activa' THEN 1 ELSE 0 END) AS activas,
                SUM(CASE WHEN a.estado = 'Finalizada' THEN 1 ELSE 0 END) AS finalizadas,
                SUM(CASE WHEN a.estado = 'Cancelada' THEN 1 ELSE 0 END) AS canceladas
            FROM inventario_asignaciones a
            INNER JOIN inventario_equipos e ON e.id = a.equipo_id
            """ + clausula_where(condiciones_asignaciones),
            tuple(parametros_asignaciones),
        )
        resumen_asignaciones = cursor.fetchone() or {}

        condiciones_responsivas = []
        parametros_responsivas = []

        if filtros.get("empresa") and filtros["empresa"] != "Todas":
            condiciones_responsivas.append("e.empresa = %s")
            parametros_responsivas.append(filtros["empresa"])

        if filtros.get("departamento") and filtros["departamento"] != "Todos":
            condiciones_responsivas.append("e.departamento = %s")
            parametros_responsivas.append(filtros["departamento"])

        condiciones_responsivas, parametros_responsivas = construir_filtro_fecha(
            filtros,
            "r.fecha_generacion",
            condiciones_responsivas,
            parametros_responsivas,
        )

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN r.estado = 'Pendiente' THEN 1 ELSE 0 END) AS pendientes,
                SUM(CASE WHEN r.estado = 'Firmada' THEN 1 ELSE 0 END) AS firmadas,
                SUM(CASE WHEN r.estado = 'Anulada' THEN 1 ELSE 0 END) AS anuladas
            FROM inventario_responsivas r
            INNER JOIN inventario_equipos e ON e.id = r.equipo_id
            """ + clausula_where(condiciones_responsivas),
            tuple(parametros_responsivas),
        )
        resumen_responsivas = cursor.fetchone() or {}

        condiciones_auditorias = []
        parametros_auditorias = []

        if filtros.get("empresa") and filtros["empresa"] != "Todas":
            condiciones_auditorias.append("a.empresa = %s")
            parametros_auditorias.append(filtros["empresa"])

        if filtros.get("departamento") and filtros["departamento"] != "Todos":
            condiciones_auditorias.append("a.departamento = %s")
            parametros_auditorias.append(filtros["departamento"])

        condiciones_auditorias, parametros_auditorias = construir_filtro_fecha(
            filtros,
            "a.fecha_programada",
            condiciones_auditorias,
            parametros_auditorias,
        )

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN a.estado = 'Planeada' THEN 1 ELSE 0 END) AS planeadas,
                SUM(CASE WHEN a.estado = 'En proceso' THEN 1 ELSE 0 END) AS en_proceso,
                SUM(CASE WHEN a.estado = 'Finalizada' THEN 1 ELSE 0 END) AS finalizadas,
                SUM(CASE WHEN a.estado = 'Cancelada' THEN 1 ELSE 0 END) AS canceladas
            FROM inventario_auditorias a
            """ + clausula_where(condiciones_auditorias),
            tuple(parametros_auditorias),
        )
        resumen_auditorias = cursor.fetchone() or {}

        cursor.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN d.resultado = 'Con diferencia' THEN 1 ELSE 0 END) AS diferencias,
                SUM(CASE WHEN d.resultado = 'No localizado' THEN 1 ELSE 0 END) AS no_localizados,
                SUM(
                    CASE
                        WHEN d.estado_correccion IN ('Pendiente', 'En proceso')
                        THEN 1
                        ELSE 0
                    END
                ) AS correcciones_pendientes
            FROM inventario_auditoria_detalles d
            INNER JOIN inventario_auditorias a ON a.id = d.auditoria_id
            """ + clausula_where(condiciones_auditorias),
            tuple(parametros_auditorias),
        )
        resumen_hallazgos = cursor.fetchone() or {}

        distribuciones = {}

        for clave, columna in [
            ("porEstado", "e.estado"),
            ("porCategoria", "COALESCE(NULLIF(TRIM(e.categoria), ''), 'Sin categoría')"),
            ("porEmpresa", "COALESCE(NULLIF(TRIM(e.empresa), ''), 'Sin empresa')"),
            ("porDepartamento", "COALESCE(NULLIF(TRIM(e.departamento), ''), 'Sin departamento')"),
            ("porFuncionamiento", "COALESCE(NULLIF(TRIM(e.funcionamiento), ''), 'Sin información')"),
            ("porResponsiva", "COALESCE(NULLIF(TRIM(e.responsiva_estado), ''), 'No aplica')"),
        ]:
            cursor.execute(
                f"""
                SELECT {columna} AS etiqueta, COUNT(*) AS total
                FROM inventario_equipos e
                {filtro_equipos}
                GROUP BY {columna}
                ORDER BY total DESC, etiqueta
                """,
                tuple(parametros_equipos),
            )
            distribuciones[clave] = [
                {
                    "etiqueta": fila.get("etiqueta") or "Sin información",
                    "total": int(fila.get("total") or 0),
                }
                for fila in cursor.fetchall()
            ]

        cursor.execute(
            """
            SELECT
                DATE_FORMAT(a.fecha_asignacion, '%Y-%m') AS periodo,
                COUNT(*) AS total
            FROM inventario_asignaciones a
            INNER JOIN inventario_equipos e ON e.id = a.equipo_id
            """ + clausula_where(condiciones_asignaciones) + """
            GROUP BY DATE_FORMAT(a.fecha_asignacion, '%Y-%m')
            ORDER BY periodo
            """,
            tuple(parametros_asignaciones),
        )
        asignaciones_por_mes = [
            {
                "periodo": fila.get("periodo") or "",
                "total": int(fila.get("total") or 0),
            }
            for fila in cursor.fetchall()
        ]

        cursor.execute(
            """
            SELECT
                DATE_FORMAT(a.fecha_devolucion, '%Y-%m') AS periodo,
                COUNT(*) AS total
            FROM inventario_asignaciones a
            INNER JOIN inventario_equipos e ON e.id = a.equipo_id
            """ + clausula_where(
                condiciones_asignaciones + ["a.fecha_devolucion IS NOT NULL"]
            ) + """
            GROUP BY DATE_FORMAT(a.fecha_devolucion, '%Y-%m')
            ORDER BY periodo
            """,
            tuple(parametros_asignaciones),
        )
        devoluciones_por_mes = [
            {
                "periodo": fila.get("periodo") or "",
                "total": int(fila.get("total") or 0),
            }
            for fila in cursor.fetchall()
        ]

        movimientos_recientes = []

        if tabla_existe(cursor, "inventario_movimientos"):
            fecha_movimiento = obtener_columna_fecha_movimientos(cursor)
            id_movimiento = obtener_columna_id_movimientos(cursor)

            if fecha_movimiento:
                orden = f"m.{fecha_movimiento} DESC"
                if id_movimiento:
                    orden += f", m.{id_movimiento} DESC"

                cursor.execute(
                    f"""
                    SELECT
                        {f'm.{id_movimiento} AS id,' if id_movimiento else 'NULL AS id,'}
                        m.tipo_movimiento,
                        m.descripcion,
                        m.responsable_anterior,
                        m.responsable_nuevo,
                        m.usuario_registro,
                        m.{fecha_movimiento} AS fecha,
                        e.numero_inventario,
                        e.descripcion AS equipo
                    FROM inventario_movimientos m
                    LEFT JOIN inventario_equipos e ON e.id = m.equipo_id
                    ORDER BY {orden}
                    LIMIT 10
                    """
                )
                movimientos_recientes = convertir_filas_json(cursor.fetchall())

        return jsonify(
            {
                "resumen": {
                    "equipos": {
                        "total": int(resumen_equipos.get("total") or 0),
                        "asignados": int(resumen_equipos.get("asignados") or 0),
                        "disponibles": int(resumen_equipos.get("disponibles") or 0),
                        "bajas": int(resumen_equipos.get("bajas") or 0),
                        "responsivasPendientes": int(
                            resumen_equipos.get("responsivas_pendientes") or 0
                        ),
                        "responsivasFirmadas": int(
                            resumen_equipos.get("responsivas_firmadas") or 0
                        ),
                    },
                    "asignaciones": {
                        "total": int(resumen_asignaciones.get("total") or 0),
                        "activas": int(resumen_asignaciones.get("activas") or 0),
                        "finalizadas": int(resumen_asignaciones.get("finalizadas") or 0),
                        "canceladas": int(resumen_asignaciones.get("canceladas") or 0),
                    },
                    "responsivas": {
                        "total": int(resumen_responsivas.get("total") or 0),
                        "pendientes": int(resumen_responsivas.get("pendientes") or 0),
                        "firmadas": int(resumen_responsivas.get("firmadas") or 0),
                        "anuladas": int(resumen_responsivas.get("anuladas") or 0),
                    },
                    "auditorias": {
                        "total": int(resumen_auditorias.get("total") or 0),
                        "planeadas": int(resumen_auditorias.get("planeadas") or 0),
                        "enProceso": int(resumen_auditorias.get("en_proceso") or 0),
                        "finalizadas": int(resumen_auditorias.get("finalizadas") or 0),
                        "canceladas": int(resumen_auditorias.get("canceladas") or 0),
                    },
                    "hallazgos": {
                        "totalRevisiones": int(resumen_hallazgos.get("total") or 0),
                        "diferencias": int(resumen_hallazgos.get("diferencias") or 0),
                        "noLocalizados": int(resumen_hallazgos.get("no_localizados") or 0),
                        "correccionesPendientes": int(
                            resumen_hallazgos.get("correcciones_pendientes") or 0
                        ),
                    },
                },
                "distribuciones": distribuciones,
                "tendencias": {
                    "asignacionesPorMes": asignaciones_por_mes,
                    "devolucionesPorMes": devoluciones_por_mes,
                },
                "movimientosRecientes": movimientos_recientes,
            }
        )

    except Exception as error:
        print("Error al generar dashboard de reportes:", error)
        return jsonify(
            {
                "error": "No se pudo generar el dashboard de reportes.",
                "detalle": str(error),
            }
        ), 500

    finally:
        cursor.close()
        conexion.close()


# ============================================================
# GET: DETALLE TABULAR POR TIPO
# ============================================================


@reportes_bp.route("/detalle/<string:tipo>", methods=["GET"])
def obtener_detalle(tipo):
    try:
        filtros = obtener_filtros()
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    tipo = tipo.strip().lower()

    if tipo not in {
        "equipos",
        "asignaciones",
        "responsivas",
        "auditorias",
        "movimientos",
    }:
        return jsonify({"error": "El tipo de reporte no es válido."}), 400

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    try:
        if tipo == "equipos":
            clausula, parametros = construir_filtro_equipos(filtros)
            cursor.execute(
                """
                SELECT
                    e.id,
                    e.numero_inventario AS inventario,
                    e.descripcion AS equipo,
                    e.categoria,
                    e.marca,
                    e.modelo,
                    e.numero_serie AS serie,
                    e.empresa,
                    e.departamento,
                    e.responsable,
                    e.estado,
                    e.funcionamiento,
                    e.ubicacion,
                    e.responsiva_estado AS responsiva
                FROM inventario_equipos e
                """ + clausula + " ORDER BY e.numero_inventario",
                tuple(parametros),
            )

        elif tipo == "asignaciones":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("a.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "a.fecha_asignacion",
                condiciones,
                parametros,
            )

            if filtros.get("busqueda"):
                termino = f"%{filtros['busqueda']}%"
                condiciones.append(
                    """
                    (
                        e.numero_inventario LIKE %s
                        OR e.descripcion LIKE %s
                        OR CONCAT_WS(' ', c.nombre, c.apellido_paterno, c.apellido_materno) LIKE %s
                        OR c.numero_empleado LIKE %s
                    )
                    """
                )
                parametros.extend([termino] * 4)

            cursor.execute(
                """
                SELECT
                    a.id,
                    e.numero_inventario AS inventario,
                    e.descripcion AS equipo,
                    CONCAT_WS(' ', c.nombre, c.apellido_paterno, c.apellido_materno) AS colaborador,
                    c.numero_empleado,
                    c.puesto,
                    e.empresa,
                    e.departamento,
                    a.fecha_asignacion,
                    a.fecha_devolucion,
                    a.estado,
                    a.observaciones_entrega,
                    a.observaciones_devolucion,
                    a.usuario_registro
                FROM inventario_asignaciones a
                INNER JOIN inventario_equipos e ON e.id = a.equipo_id
                INNER JOIN inventario_colaboradores c ON c.id = a.colaborador_id
                """ + clausula_where(condiciones) + """
                ORDER BY a.fecha_asignacion DESC, a.id DESC
                """,
                tuple(parametros),
            )

        elif tipo == "responsivas":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("r.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "r.fecha_generacion",
                condiciones,
                parametros,
            )

            if filtros.get("busqueda"):
                termino = f"%{filtros['busqueda']}%"
                condiciones.append(
                    """
                    (
                        r.folio LIKE %s
                        OR e.numero_inventario LIKE %s
                        OR e.descripcion LIKE %s
                        OR r.responsable LIKE %s
                    )
                    """
                )
                parametros.extend([termino] * 4)

            cursor.execute(
                """
                SELECT
                    r.id,
                    r.folio,
                    e.numero_inventario AS inventario,
                    e.descripcion AS equipo,
                    r.responsable,
                    e.empresa,
                    e.departamento,
                    r.estado,
                    r.fecha_generacion,
                    r.fecha_firma,
                    r.fecha_anulacion,
                    r.motivo_anulacion,
                    r.observaciones,
                    r.archivo_pdf
                FROM inventario_responsivas r
                INNER JOIN inventario_equipos e ON e.id = r.equipo_id
                """ + clausula_where(condiciones) + """
                ORDER BY r.fecha_generacion DESC, r.id DESC
                """,
                tuple(parametros),
            )

        elif tipo == "auditorias":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("a.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("a.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("a.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "a.fecha_programada",
                condiciones,
                parametros,
            )

            if filtros.get("busqueda"):
                termino = f"%{filtros['busqueda']}%"
                condiciones.append(
                    """
                    (
                        a.folio LIKE %s
                        OR a.nombre LIKE %s
                        OR a.auditor_responsable LIKE %s
                    )
                    """
                )
                parametros.extend([termino] * 3)

            cursor.execute(
                """
                SELECT
                    a.id,
                    a.folio,
                    a.nombre,
                    a.tipo,
                    a.empresa,
                    a.departamento,
                    a.ubicacion,
                    a.fecha_programada,
                    a.fecha_inicio,
                    a.fecha_finalizacion,
                    a.estado,
                    a.auditor_responsable,
                    a.observaciones,
                    a.conclusiones,
                    COALESCE(r.total_equipos, 0) AS total_equipos,
                    COALESCE(r.revisados, 0) AS revisados,
                    COALESCE(r.diferencias, 0) AS diferencias,
                    COALESCE(r.no_localizados, 0) AS no_localizados,
                    COALESCE(r.correcciones_pendientes, 0) AS correcciones_pendientes
                FROM inventario_auditorias a
                LEFT JOIN (
                    SELECT
                        auditoria_id,
                        COUNT(*) AS total_equipos,
                        SUM(CASE WHEN resultado <> 'Pendiente' THEN 1 ELSE 0 END) AS revisados,
                        SUM(CASE WHEN resultado = 'Con diferencia' THEN 1 ELSE 0 END) AS diferencias,
                        SUM(CASE WHEN resultado = 'No localizado' THEN 1 ELSE 0 END) AS no_localizados,
                        SUM(
                            CASE
                                WHEN estado_correccion IN ('Pendiente', 'En proceso')
                                THEN 1
                                ELSE 0
                            END
                        ) AS correcciones_pendientes
                    FROM inventario_auditoria_detalles
                    GROUP BY auditoria_id
                ) r ON r.auditoria_id = a.id
                """ + clausula_where(condiciones) + """
                ORDER BY a.fecha_programada DESC, a.id DESC
                """,
                tuple(parametros),
            )

        else:
            if not tabla_existe(cursor, "inventario_movimientos"):
                return jsonify([])

            fecha_movimiento = obtener_columna_fecha_movimientos(cursor)
            id_movimiento = obtener_columna_id_movimientos(cursor)

            if not fecha_movimiento:
                return jsonify(
                    {
                        "error": "La tabla de movimientos no tiene una columna de fecha reconocida."
                    }
                ), 409

            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                f"m.{fecha_movimiento}",
                condiciones,
                parametros,
            )

            if filtros.get("busqueda"):
                termino = f"%{filtros['busqueda']}%"
                condiciones.append(
                    """
                    (
                        e.numero_inventario LIKE %s
                        OR e.descripcion LIKE %s
                        OR m.tipo_movimiento LIKE %s
                        OR m.descripcion LIKE %s
                        OR m.usuario_registro LIKE %s
                    )
                    """
                )
                parametros.extend([termino] * 5)

            orden = f"m.{fecha_movimiento} DESC"
            if id_movimiento:
                orden += f", m.{id_movimiento} DESC"

            cursor.execute(
                f"""
                SELECT
                    {f'm.{id_movimiento} AS id,' if id_movimiento else 'NULL AS id,'}
                    e.numero_inventario AS inventario,
                    e.descripcion AS equipo,
                    e.empresa,
                    e.departamento,
                    m.tipo_movimiento,
                    m.descripcion,
                    m.responsable_anterior,
                    m.responsable_nuevo,
                    m.usuario_registro,
                    m.{fecha_movimiento} AS fecha
                FROM inventario_movimientos m
                LEFT JOIN inventario_equipos e ON e.id = m.equipo_id
                {clausula_where(condiciones)}
                ORDER BY {orden}
                """,
                tuple(parametros),
            )

        return jsonify(convertir_filas_json(cursor.fetchall()))

    except Exception as error:
        print(f"Error al obtener reporte {tipo}:", error)
        return jsonify(
            {
                "error": f"No se pudo obtener el reporte de {tipo}.",
                "detalle": str(error),
            }
        ), 500

    finally:
        cursor.close()
        conexion.close()


# ============================================================
# GET: EXPORTAR REPORTE A EXCEL
# ============================================================


@reportes_bp.route("/exportar/<string:tipo>", methods=["GET"])
def exportar_reporte(tipo):
    tipo = tipo.strip().lower()

    if tipo not in {
        "equipos",
        "asignaciones",
        "responsivas",
        "auditorias",
        "movimientos",
    }:
        return jsonify({"error": "El tipo de reporte no es válido."}), 400

    try:
        filtros = obtener_filtros()
    except ValueError as error:
        return jsonify({"error": str(error)}), 400

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    try:
        # Consulta directa por tipo para conservar nombres amigables en Excel.
        if tipo == "equipos":
            clausula, parametros = construir_filtro_equipos(filtros)
            cursor.execute(
                """
                SELECT
                    e.numero_inventario,
                    e.descripcion,
                    e.categoria,
                    e.marca,
                    e.modelo,
                    e.numero_serie,
                    e.empresa,
                    e.departamento,
                    e.responsable,
                    e.estado,
                    e.funcionamiento,
                    e.ubicacion,
                    e.responsiva_estado
                FROM inventario_equipos e
                """ + clausula + " ORDER BY e.numero_inventario",
                tuple(parametros),
            )
            encabezados = [
                "No. inventario",
                "Equipo",
                "Categoría",
                "Marca",
                "Modelo",
                "Número de serie",
                "Empresa",
                "Departamento",
                "Responsable",
                "Estado",
                "Funcionamiento",
                "Ubicación",
                "Responsiva",
            ]

        elif tipo == "asignaciones":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("a.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "a.fecha_asignacion",
                condiciones,
                parametros,
            )

            cursor.execute(
                """
                SELECT
                    e.numero_inventario,
                    e.descripcion,
                    CONCAT_WS(' ', c.nombre, c.apellido_paterno, c.apellido_materno) AS colaborador,
                    c.numero_empleado,
                    c.puesto,
                    e.empresa,
                    e.departamento,
                    a.fecha_asignacion,
                    a.fecha_devolucion,
                    a.estado,
                    a.observaciones_entrega,
                    a.observaciones_devolucion,
                    a.usuario_registro
                FROM inventario_asignaciones a
                INNER JOIN inventario_equipos e ON e.id = a.equipo_id
                INNER JOIN inventario_colaboradores c ON c.id = a.colaborador_id
                """ + clausula_where(condiciones) + """
                ORDER BY a.fecha_asignacion DESC, a.id DESC
                """,
                tuple(parametros),
            )
            encabezados = [
                "No. inventario",
                "Equipo",
                "Colaborador",
                "No. empleado",
                "Puesto",
                "Empresa",
                "Departamento",
                "Fecha asignación",
                "Fecha devolución",
                "Estado",
                "Observaciones entrega",
                "Observaciones devolución",
                "Registró",
            ]

        elif tipo == "responsivas":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("r.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "r.fecha_generacion",
                condiciones,
                parametros,
            )

            cursor.execute(
                """
                SELECT
                    r.folio,
                    e.numero_inventario,
                    e.descripcion,
                    r.responsable,
                    e.empresa,
                    e.departamento,
                    r.estado,
                    r.fecha_generacion,
                    r.fecha_firma,
                    r.fecha_anulacion,
                    r.motivo_anulacion,
                    r.observaciones,
                    r.archivo_pdf
                FROM inventario_responsivas r
                INNER JOIN inventario_equipos e ON e.id = r.equipo_id
                """ + clausula_where(condiciones) + """
                ORDER BY r.fecha_generacion DESC, r.id DESC
                """,
                tuple(parametros),
            )
            encabezados = [
                "Folio",
                "No. inventario",
                "Equipo",
                "Responsable",
                "Empresa",
                "Departamento",
                "Estado",
                "Fecha generación",
                "Fecha firma",
                "Fecha anulación",
                "Motivo anulación",
                "Observaciones",
                "Documento firmado",
            ]

        elif tipo == "auditorias":
            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("a.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("a.departamento = %s")
                parametros.append(filtros["departamento"])

            if filtros.get("estado") and filtros["estado"] != "Todos":
                condiciones.append("a.estado = %s")
                parametros.append(filtros["estado"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                "a.fecha_programada",
                condiciones,
                parametros,
            )

            cursor.execute(
                """
                SELECT
                    a.folio,
                    a.nombre,
                    a.tipo,
                    a.empresa,
                    a.departamento,
                    a.ubicacion,
                    a.fecha_programada,
                    a.fecha_inicio,
                    a.fecha_finalizacion,
                    a.estado,
                    a.auditor_responsable,
                    a.observaciones,
                    a.conclusiones,
                    COALESCE(r.total_equipos, 0) AS total_equipos,
                    COALESCE(r.revisados, 0) AS revisados,
                    COALESCE(r.diferencias, 0) AS diferencias,
                    COALESCE(r.no_localizados, 0) AS no_localizados,
                    COALESCE(r.correcciones_pendientes, 0) AS correcciones_pendientes
                FROM inventario_auditorias a
                LEFT JOIN (
                    SELECT
                        auditoria_id,
                        COUNT(*) AS total_equipos,
                        SUM(CASE WHEN resultado <> 'Pendiente' THEN 1 ELSE 0 END) AS revisados,
                        SUM(CASE WHEN resultado = 'Con diferencia' THEN 1 ELSE 0 END) AS diferencias,
                        SUM(CASE WHEN resultado = 'No localizado' THEN 1 ELSE 0 END) AS no_localizados,
                        SUM(
                            CASE
                                WHEN estado_correccion IN ('Pendiente', 'En proceso')
                                THEN 1
                                ELSE 0
                            END
                        ) AS correcciones_pendientes
                    FROM inventario_auditoria_detalles
                    GROUP BY auditoria_id
                ) r ON r.auditoria_id = a.id
                """ + clausula_where(condiciones) + """
                ORDER BY a.fecha_programada DESC, a.id DESC
                """,
                tuple(parametros),
            )
            encabezados = [
                "Folio",
                "Nombre",
                "Tipo",
                "Empresa",
                "Departamento",
                "Ubicación",
                "Fecha programada",
                "Fecha inicio",
                "Fecha finalización",
                "Estado",
                "Auditor",
                "Observaciones",
                "Conclusiones",
                "Total equipos",
                "Revisados",
                "Diferencias",
                "No localizados",
                "Correcciones pendientes",
            ]

        else:
            if not tabla_existe(cursor, "inventario_movimientos"):
                return jsonify({"error": "No existe la tabla de movimientos."}), 404

            fecha_movimiento = obtener_columna_fecha_movimientos(cursor)
            id_movimiento = obtener_columna_id_movimientos(cursor)

            if not fecha_movimiento:
                return jsonify(
                    {
                        "error": "La tabla de movimientos no tiene una columna de fecha reconocida."
                    }
                ), 409

            condiciones = []
            parametros = []

            if filtros.get("empresa") and filtros["empresa"] != "Todas":
                condiciones.append("e.empresa = %s")
                parametros.append(filtros["empresa"])

            if filtros.get("departamento") and filtros["departamento"] != "Todos":
                condiciones.append("e.departamento = %s")
                parametros.append(filtros["departamento"])

            condiciones, parametros = construir_filtro_fecha(
                filtros,
                f"m.{fecha_movimiento}",
                condiciones,
                parametros,
            )

            orden = f"m.{fecha_movimiento} DESC"
            if id_movimiento:
                orden += f", m.{id_movimiento} DESC"

            cursor.execute(
                f"""
                SELECT
                    e.numero_inventario,
                    e.descripcion,
                    e.empresa,
                    e.departamento,
                    m.tipo_movimiento,
                    m.descripcion,
                    m.responsable_anterior,
                    m.responsable_nuevo,
                    m.usuario_registro,
                    m.{fecha_movimiento}
                FROM inventario_movimientos m
                LEFT JOIN inventario_equipos e ON e.id = m.equipo_id
                {clausula_where(condiciones)}
                ORDER BY {orden}
                """,
                tuple(parametros),
            )
            encabezados = [
                "No. inventario",
                "Equipo",
                "Empresa",
                "Departamento",
                "Tipo de movimiento",
                "Descripción",
                "Responsable anterior",
                "Responsable nuevo",
                "Registró",
                "Fecha",
            ]

        registros = cursor.fetchall()
        filas = []

        for registro in registros:
            filas.append(
                [
                    fecha_a_texto(valor)
                    if isinstance(valor, (date, datetime))
                    else valor
                    for valor in registro.values()
                ]
            )

        libro = Workbook()
        hoja = libro.active
        preparar_hoja(
            hoja,
            tipo.capitalize(),
            encabezados,
            filas,
        )

        buffer = BytesIO()
        libro.save(buffer)
        buffer.seek(0)

        fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_{tipo}_{fecha_archivo}.xlsx"

        return send_file(
            buffer,
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=nombre_archivo,
        )

    except Exception as error:
        print(f"Error al exportar reporte {tipo}:", error)
        return jsonify(
            {
                "error": f"No se pudo exportar el reporte de {tipo}.",
                "detalle": str(error),
            }
        ), 500

    finally:
        cursor.close()
        conexion.close()