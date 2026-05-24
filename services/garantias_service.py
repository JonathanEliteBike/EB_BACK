from __future__ import annotations
import io
import json
import logging

from utils.tiempo import ahora_str


def invalidar_cache() -> None:
    """No-op: el dashboard ahora lee directo de BD, no hay caché que invalidar."""
    pass


def get_dashboard_data() -> dict:
    """Calcula KPIs y métricas desde garantia_formularios (BD MySQL)."""
    from db_conexion import obtener_conexion

    conn = obtener_conexion()
    if not conn:
        return _empty()
    try:
        cursor = conn.cursor(dictionary=True)

        # ── KPIs principales ─────────────────────────────────────────────────
        cursor.execute("""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN estatus = 'Cerrado' THEN 1 ELSE 0 END) AS cerradas,
                SUM(CASE WHEN estatus IN ('Enviado','En revisión','Aprobado') THEN 1 ELSE 0 END) AS en_proceso
            FROM garantia_formularios
        """)
        kpi = cursor.fetchone() or {}

        # Latencia de atención: días desde creación hasta primer evento de validación de docs
        cursor.execute("""
            SELECT ROUND(AVG(dias), 1) AS lat_atencion FROM (
                SELECT f.id, DATEDIFF(MIN(c.fecha), f.fecha_creacion) AS dias
                FROM garantia_formularios f
                JOIN garantia_comentarios c ON c.formulario_id = f.id
                     AND c.tipo = 'validacion'
                GROUP BY f.id
                HAVING dias >= 0
            ) t
        """)
        lat_atencion = float((cursor.fetchone() or {}).get('lat_atencion') or 0)

        # Latencia de cierre: días desde creación hasta que el ticket fue Cerrado o Rechazado
        cursor.execute("""
            SELECT ROUND(AVG(DATEDIFF(fecha_actualizacion, fecha_creacion)), 1) AS lat_cierre
            FROM garantia_formularios
            WHERE estatus IN ('Cerrado', 'Rechazado')
        """)
        lat_cierre = float((cursor.fetchone() or {}).get('lat_cierre') or 0)

        # ── Por estatus ──────────────────────────────────────────────────────
        cursor.execute("""
            SELECT estatus, COUNT(*) AS cnt
            FROM garantia_formularios
            GROUP BY estatus ORDER BY cnt DESC
        """)
        por_estatus = {r['estatus']: r['cnt'] for r in cursor.fetchall()}

        # ── Tickets por mes (latencia = días medios de antigüedad ese mes) ───
        cursor.execute("""
            SELECT DATE_FORMAT(fecha_creacion, '%Y-%m') AS ym,
                   DATE_FORMAT(fecha_creacion, '%b %Y')  AS label,
                   ROUND(AVG(DATEDIFF(NOW(), fecha_creacion)), 1) AS lat_prom
            FROM garantia_formularios
            GROUP BY ym, label
            ORDER BY ym
        """)
        latencia_mensual = {r['label']: r['lat_prom'] for r in cursor.fetchall()}

        # ── Garantías por distribuidor (top 30) ──────────────────────────────
        cursor.execute("""
            SELECT distribuidor AS cli, COUNT(*) AS cnt
            FROM garantia_formularios
            WHERE distribuidor IS NOT NULL AND distribuidor != ''
            GROUP BY distribuidor ORDER BY cnt DESC LIMIT 30
        """)
        gar_cliente = {r['cli']: r['cnt'] for r in cursor.fetchall()}

        # ── Latencia promedio por distribuidor (top 30) ──────────────────────
        cursor.execute("""
            SELECT distribuidor AS cli,
                   ROUND(AVG(DATEDIFF(NOW(), fecha_creacion)), 1) AS lat
            FROM garantia_formularios
            WHERE distribuidor IS NOT NULL AND distribuidor != ''
            GROUP BY distribuidor ORDER BY lat DESC LIMIT 30
        """)
        lat_cliente = {r['cli']: r['lat'] for r in cursor.fetchall()}

        # ── Por marca ────────────────────────────────────────────────────────
        cursor.execute("""
            SELECT marca, COUNT(*) AS cnt
            FROM garantia_formularios
            WHERE marca IS NOT NULL AND marca != ''
            GROUP BY marca ORDER BY cnt DESC
        """)
        por_marca = {r['marca']: r['cnt'] for r in cursor.fetchall()}

        # ── Piezas de reemplazo (columna directa) ────────────────────────────
        cursor.execute("""
            SELECT pieza_reemplazo AS pieza, COUNT(*) AS cnt
            FROM garantia_formularios
            WHERE pieza_reemplazo IS NOT NULL
              AND pieza_reemplazo != ''
              AND pieza_reemplazo != 'N/A'
            GROUP BY pieza_reemplazo ORDER BY cnt DESC
        """)
        piezas_reemplazo = {r['pieza']: r['cnt'] for r in cursor.fetchall()}

        # ── Ubicación del daño (desde JSON datos) ────────────────────────────
        cursor.execute("SELECT datos FROM garantia_formularios WHERE datos IS NOT NULL")
        ubic_dano: dict = {}
        for row in cursor.fetchall():
            try:
                d = json.loads(row['datos']) if isinstance(row['datos'], str) else (row['datos'] or {})
                for k, v in d.items():
                    if k.startswith('marco_localizacion_') and v:
                        key = str(v).strip()
                        ubic_dano[key] = ubic_dano.get(key, 0) + 1
            except Exception:
                pass
        ubic_dano = dict(sorted(ubic_dano.items(), key=lambda x: x[1], reverse=True)[:25])

        return {
            "kpis": {
                "total":        int(kpi.get("total", 0) or 0),
                "cerradas":     int(kpi.get("cerradas", 0) or 0),
                "en_proceso":   int(kpi.get("en_proceso", 0) or 0),
                "lat_atencion": lat_atencion,
                "lat_cierre":   lat_cierre,
            },
            "por_estatus":           por_estatus,
            "latencia_mensual":      latencia_mensual,
            "latencia_por_cliente":  lat_cliente,
            "garantias_por_cliente": gar_cliente,
            "piezas_reemplazo":      piezas_reemplazo,
            "ubicacion_dano":        ubic_dano,
            "por_marca":             por_marca,
            "ultima_actualizacion":  ahora_str("%d/%m/%Y %H:%M"),
        }
    except Exception as e:
        logging.exception("Error en get_dashboard_data (DB): %s", e)
        return _empty()
    finally:
        conn.close()


def _empty() -> dict:
    return {
        "kpis": {"total": 0, "cerradas": 0, "en_proceso": 0, "lat_atencion": 0.0, "lat_cierre": 0.0},
        "por_estatus": {}, "latencia_mensual": {}, "latencia_por_cliente": {},
        "garantias_por_cliente": {}, "piezas_reemplazo": {}, "ubicacion_dano": {},
        "por_marca": {}, "ultima_actualizacion": ahora_str("%d/%m/%Y %H:%M"),
    }


def exportar_excel() -> bytes:
    """Exporta todos los tickets de garantía a Excel con columnas completas."""
    try:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill
        from db_conexion import obtener_conexion

        conn = obtener_conexion()
        if not conn:
            return b""
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                folio,
                distribuidor,
                contacto,
                puesto,
                email,
                marca,
                estatus,
                estatus_pieza,
                pieza_reemplazo,
                docs_validados,
                serie_validada,
                DATE_FORMAT(fecha_creacion,    '%%d/%%m/%%Y %%H:%%i') AS fecha_creacion,
                DATE_FORMAT(fecha_actualizacion,'%%d/%%m/%%Y %%H:%%i') AS fecha_actualizacion
            FROM garantia_formularios
            ORDER BY fecha_creacion DESC
        """)
        rows = cursor.fetchall()
        conn.close()

        # Renombrar columnas a español legible
        columnas_es = {
            "folio":               "Folio",
            "distribuidor":        "Distribuidor",
            "contacto":            "Contacto",
            "puesto":              "Puesto",
            "email":               "Correo",
            "marca":               "Marca",
            "estatus":             "Estatus",
            "estatus_pieza":       "Estado de Pieza",
            "pieza_reemplazo":     "Pieza de Reemplazo",
            "docs_validados":      "Docs Validados",
            "serie_validada":      "Serie Validada",
            "fecha_creacion":      "Fecha de Envío",
            "fecha_actualizacion": "Última Actualización",
        }

        df = pd.DataFrame(rows).rename(columns=columnas_es)

        # Convertir booleanos 0/1 a Sí/No
        for col in ["Docs Validados", "Serie Validada"]:
            if col in df.columns:
                df[col] = df[col].map({1: "Sí", 0: "No", None: "—"}).fillna("—")

        # Valores vacíos → guión
        df = df.fillna("—")

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Garantías")
            ws = writer.sheets["Garantías"]

            # Estilos de encabezado
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill("solid", fgColor="EB5E28")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

            for cell in ws[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center

            # Alineación y ancho de columnas
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                header_val = str(col_cells[0].value or "")
                max_len = max(
                    (len(str(c.value or "")) for c in col_cells),
                    default=len(header_val)
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
                for cell in col_cells[1:]:
                    cell.alignment = left

            ws.freeze_panes   = "A2"
            ws.row_dimensions[1].height = 28

        return buf.getvalue()
    except Exception as e:
        logging.exception("Error exportando Excel: %s", e)
        return b""
